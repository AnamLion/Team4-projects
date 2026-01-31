# storage.py
# ------------------------------------------------------------
# Central Excel utility module for the Library Management System.
#
# Responsibilities:
#   - Create library.xlsx if it does not exist
#   - Ensure required sheets exist
#   - Provide a single workbook instance to all modules
# ------------------------------------------------------------

import os
from openpyxl import Workbook, load_workbook

# -------- CONFIGURATION --------
LIBRARY_FILE = "library.xlsx"

REQUIRED_SHEETS = {
    "Books": [
        "BookID",
        "Title",
        "Author",
        "Quantity"
    ],
    "Members": [
        "MemberID",
        "Name",
        "Phone",
        "BooksIssued"
    ],
    "Transactions": [
        "TransactionID",
        "MemberID",
        "BookID",
        "IssueDate",
        "ReturnDate",
        "Fine"
    ]
}
# -------------------------------


class LibraryStorage:
    def __init__(self, filename="library.xlsx"):
        self.filename = filename
        self.required_sheets = {
            "Books": ["BookID", "Title", "Author", "Quantity"],
            "Members": ["MemberID", "Name", "Phone", "BooksIssued"],
            "Transactions": ["TransactionID", "MemberID", "BookID", "IssueDate", "ReturnDate", "Fine"]
        }

    def create_library_file(self):
        """Creates the Excel file if it does NOT exist and initializes sheets."""
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        for sheet_name, headers in self.required_sheets.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        wb.save(self.filename)

    def ensure_sheets_exist(self, wb):
        """Ensures that all required sheets exist inside the workbook."""
        for sheet_name, headers in self.required_sheets.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)

    def get_workbook(self):
        """Returns a single, consistent workbook instance."""
        if not os.path.exists(self.filename):
            self.create_library_file()
        wb = load_workbook(self.filename)
        self.ensure_sheets_exist(wb)
        return wb

    def save_workbook(self, wb):
        """Saves the given workbook to disk."""
        wb.save(self.filename)

    def get_sheet(self, wb, sheet_name):
        """Returns a specific worksheet from the given workbook."""
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(self.required_sheets[sheet_name])
        return wb[sheet_name]

