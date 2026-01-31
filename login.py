# login.py
import os
from openpyxl import Workbook, load_workbook

class AuthManager:
    def __init__(self, user_file="users.xlsx"):
        self.user_file = user_file
        self.sheet_name = "Users"

    def create_default_users(self):
        """Creates users.xlsx with default accounts if it does NOT exist."""
        if not os.path.exists(self.user_file):
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            ws.append(["Username", "Password", "Role"])
            ws.append(["admin", "admin123", "ADMIN"])
            ws.append(["librarian", "lib123", "LIBRARIAN"])
            wb.save(self.user_file)

    def authenticate(self, username, password):
        """Checks credentials against users.xlsx."""
        if not os.path.exists(self.user_file):
            self.create_default_users()
            
        wb = load_workbook(self.user_file)
        ws = wb[self.sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            stored_user, stored_pass, role = row
            if username == stored_user and password == stored_pass:
                return role
        return None

    def login_menu(self):
        """Displays the login menu and handles user input."""
        self.create_default_users()
        while True:
            print("LIBRARY MANAGEMENT SYSTEM")
            print("1. Login")
            print("2. Exit")

            choice = input("Enter choice: ")
            if choice == "1":
                username = input("Username: ")
                password = input("Password: ")
                role = self.authenticate(username, password)
                if role:
                    print(f"\n[SUCCESS] Welcome, {username}! Role: {role}\n")
                    return role
                else:
                    print("\n[ERROR] Invalid credentials. Please try again.\n")
            elif choice == "2":
                return None
            else:
                print("\n[ERROR] Invalid choice. Please try again.\n")
