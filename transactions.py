# transactions.py
import os
from datetime import datetime
from openpyxl import load_workbook
from models import Transaction

class TransactionManager:
    def __init__(self, storage):
        self.storage = storage
        self.sheet_name = "Transactions"
        self.fine_per_day = 10
        self.due_days = 7

    def generate_transaction_id(self, ws):
        """Generates a new incremental Transaction ID."""
        last_row = ws.max_row
        if last_row <= 1:
            return 1
        last_id = ws.cell(row=last_row, column=1).value
        try:
            return int(last_id) + 1
        except (TypeError, ValueError):
            return 1

    def calculate_fine(self, issue_date_str, return_date_str):
        """Calculates fine based on due days and daily fine rate."""
        try:
            issue_date = datetime.strptime(issue_date_str, "%Y-%m-%d")
            return_date = datetime.strptime(return_date_str, "%Y-%m-%d")
            days_diff = (return_date - issue_date).days
            if days_diff > self.due_days:
                late_days = days_diff - self.due_days
                return late_days * self.fine_per_day
            return 0
        except (ValueError, TypeError):
            return 0

    def issue_book(self):
        """Issues a book to a member and records the transaction."""
        member_id = input("Enter Member ID: ")
        book_id = input("Enter Book ID: ")

        wb = self.storage.get_workbook()
        ws_trans = self.storage.get_sheet(wb, self.sheet_name)
        
        # Check if Member exists
        if "Members" in wb.sheetnames:
            ws_members = wb["Members"]
            member_found = False
            for row in ws_members.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                if str(row[0]) == member_id:
                    member_found = True
                    break
            if not member_found:
                print(f"\n[ERROR] Member ID {member_id} not found!")
                return

        # Check if Book exists AND has Stock
        if "Books" in wb.sheetnames:
            ws_books = wb["Books"]
            book_found = False
            book_row_obj = None
            for row in ws_books.iter_rows(min_row=2):
                if not any(cell.value for cell in row): continue
                stored_id = row[0].value
                quantity = row[3].value
                if str(stored_id) == book_id:
                    book_found = True
                    if quantity > 0:
                        book_row_obj = row
                    else:
                        print(f"\n[ERROR] Book ID {book_id} is out of stock!")
                        return
                    break
            if not book_found:
                print(f"\n[ERROR] Book ID {book_id} not found!")
                return
        
        # Decrease Stock
        current_qty = book_row_obj[3].value
        book_row_obj[3].value = current_qty - 1

        # Issue the Book
        transaction_id = self.generate_transaction_id(ws_trans)
        issue_date = datetime.today().strftime("%Y-%m-%d")

        ws_trans.append([
            transaction_id,
            member_id,
            book_id,
            issue_date,
            "",     # ReturnDate empty
            0       # Fine 0
        ])

        try:
            self.storage.save_workbook(wb)
            print("\n[SUCCESS] Book issued successfully.")
            print(f"Transaction ID: {transaction_id}")
            print(f"Issue Date: {issue_date}\n")
        except PermissionError:
            print("\n[ERROR] Close the Excel file 'library.xlsx' and try again!")

    def return_book(self):
        """Returns a book by locating matching active transaction."""
        member_id = input("Enter Member ID: ")
        book_id = input("Enter Book ID: ")

        wb = self.storage.get_workbook()
        ws = self.storage.get_sheet(wb, self.sheet_name)

        found = False
        return_date = datetime.today().strftime("%Y-%m-%d")

        for row in ws.iter_rows(min_row=2):
            if not any(cell.value for cell in row): continue
            stored_member = str(row[1].value)
            stored_book = str(row[2].value)
            issue_date = row[3].value
            return_date_cell = row[4].value

            if (stored_member == member_id and stored_book == book_id and 
                (return_date_cell is None or return_date_cell == "")):
                
                fine = self.calculate_fine(issue_date, return_date)
                row[4].value = return_date
                row[5].value = fine

                # Update stock
                if "Books" in wb.sheetnames:
                    ws_books = wb["Books"]
                    for b_row in ws_books.iter_rows(min_row=2):
                        if str(b_row[0].value) == book_id:
                            b_row[3].value = (b_row[3].value or 0) + 1
                            break
                found = True
                self.storage.save_workbook(wb)
                print("\n[SUCCESS] Book returned successfully.")
                print(f"Return Date: {return_date}")
                print(f"Fine: Rs. {fine}\n")
                break

        if not found:
            print("\n[INFO] No matching active transaction found.\n")

    def transactions_menu(self):
        """Displays Transactions menu."""
        while True:
            print("TRANSACTIONS MENU")
            print("1. Issue Book")
            print("2. Return Book")
            print("3. Back to Main Menu")

            choice = input("Enter choice: ")
            if choice == "1":
                self.issue_book()
            elif choice == "2":
                self.return_book()
            elif choice == "3":
                break
            else:
                print("\n[ERROR] Invalid choice. Try again.\n")
